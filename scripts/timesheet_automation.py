"""
VCC Timesheet Automation System
================================
Full end-to-end automation for fortnightly timesheet consolidation.

Steps:
1. Completion check — ping missing via Teams
2. Download/consolidate timesheets
3. Generate Excel per project with standard format
4. Create summary sheet
5. Validate (hours, duplicates, leaves)
6. Send to Anand for review
7. After approval, email clients
8. CC finance for invoicing

Usage:
  python timesheet_automation.py check       # Step 1: Check completion, ping missing
  python timesheet_automation.py generate    # Steps 2-5: Generate consolidated Excel
  python timesheet_automation.py send-anand  # Step 6: Email to Anand for review
  python timesheet_automation.py edit        # Step 6b: Manual edit from Anand's feedback
  python timesheet_automation.py ai-review   # Step 6b: AI-powered parse of Anand's reply + auto-edit/ping
  python timesheet_automation.py ai-resolve  # Step 6c: Check Teams replies & resolve pending items
  python timesheet_automation.py send-final   # Step 7: Send approved timesheets to Gayathri
  python timesheet_automation.py leave-scan  # Scan mailbox for leave notifications
  python timesheet_automation.py leave-status # Show leave summary
  python timesheet_automation.py leave-adjust # Deduct leave hours from timesheets
  python timesheet_automation.py status      # Show current period status + leaves

  # Hiring Pipeline
  python timesheet_automation.py hire-watch         # AUTO MODE: Watch mailbox, trigger full pipeline
  python timesheet_automation.py hire-scan          # Scan emails for requirements + profiles
  python timesheet_automation.py hire-jd <req_id>   # AI-generate Job Description
  python timesheet_automation.py hire-profiles      # Scan for new profile submissions
  python timesheet_automation.py hire-status        # Show hiring pipeline dashboard
  python timesheet_automation.py hire-onboard <name> # Send onboarding email for new hire
  python timesheet_automation.py hire-update <id> <status> # Update profile pipeline status
  python timesheet_automation.py hire-bench         # Check bench resources + laptop reminders
"""

import os
import sys
import json
import time
import urllib.request
import urllib.parse
import urllib.error
import ssl
from datetime import datetime, timedelta
from pathlib import Path

# Try importing openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

# Supabase for dashboard sync.
# SUPABASE_KEY is a service-role key — it bypasses RLS entirely. Never hardcode
# it; export it in the environment of whatever runs this script.
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

_supabase_client = None
def get_supabase():
    global _supabase_client
    if _supabase_client is None:
        if not SUPABASE_URL or not SUPABASE_KEY:
            print("  WARNING: SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set. Dashboard sync disabled.")
            return None
        try:
            from supabase import create_client
            _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
        except ImportError:
            print("  WARNING: supabase-py not installed. Dashboard sync disabled.")
            return None
    return _supabase_client

# ═══════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════

# Microsoft 365 app registration. All three come from the environment — the
# client secret in particular must never be committed.
TENANT_ID = os.environ.get("AZURE_TENANT_ID")
CLIENT_ID = os.environ.get("AZURE_CLIENT_ID")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET")

if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
    print(
        "ERROR: Microsoft 365 credentials not set. Export them first:\n"
        "  export AZURE_TENANT_ID=...\n"
        "  export AZURE_CLIENT_ID=...\n"
        "  export AZURE_CLIENT_SECRET=..."
    )
    sys.exit(1)

SENDER_EMAIL = "sysadmin@d4insight.com"

# ── TESTING MODE ──────────────────────────────────────────────
# All recipient emails redirect to TEST_EMAIL during testing.
# Set TEST_MODE = False and restore real emails for production.
TEST_MODE = True
TEST_EMAIL = "gayathri.m@d4insight.com"

if TEST_MODE:
    ANAND_EMAIL = TEST_EMAIL
    KISHAN_EMAIL = TEST_EMAIL
    NICK_EMAIL = TEST_EMAIL
    AR_EMAIL = TEST_EMAIL
else:
    ANAND_EMAIL = "anand.pandy@d4insight.com"
    KISHAN_EMAIL = "kishan@d4insight.com"
    NICK_EMAIL = "nbonacci@visualcomfort.com"
    AR_EMAIL = "ar@d4insight.com"

# Leave tracker email - anyone CC'ing this is marking leave
# Using Gayathri's email as placeholder until dedicated mailbox is set up
LEAVE_TRACKER_EMAIL = "sysadmin@d4insight.com"

# Leave thresholds (per fortnight period = 10 working days)
LEAVE_THRESHOLD_WARNING = 3   # 3+ days → notify manager
LEAVE_THRESHOLD_CRITICAL = 5  # 5+ days → escalate to Anand

# Manager mapping: resource name → manager email
RESOURCE_MANAGERS = {
    # IT_ALL
    "Senthilnathan": "kishan@d4insight.com",
    "Vivekanandan": "kishan@d4insight.com",
    "Aldrin": "kishan@d4insight.com",
    "Marcos": "kishan@d4insight.com",
    "Janani": "kishan@d4insight.com",
    "Janna": "kishan@d4insight.com",
    "Karthikeyan": "kishan@d4insight.com",
    "Mohammad": "kishan@d4insight.com",
    "Javal": "kishan@d4insight.com",
    # D365_FO
    "Manishkumar": "kishan@d4insight.com",
    "Bradley": "kishan@d4insight.com",
    "Kishore": "kishan@d4insight.com",
    "Krupa": "kishan@d4insight.com",
    "Andrea": "kishan@d4insight.com",
    "Sankar": "kishan@d4insight.com",
    "Anant": "kishan@d4insight.com",
    "Shahul": "kishan@d4insight.com",
    "Boyanarasimha": "kishan@d4insight.com",
    "Srinivasan": "kishan@d4insight.com",
    "Aravindh": "kishan@d4insight.com",
    "Abhinandhan": "kishan@d4insight.com",
    "Meenalochini": "kishan@d4insight.com",
    "Keerthivasan": "kishan@d4insight.com",
    # Default for everyone else
}
DEFAULT_MANAGER = "kishan@d4insight.com"

LEAVE_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output", "leaves")

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output", "timesheets")

ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

# ═══════════════════════════════════════════════════════════════
# PROJECT CONFIGURATION - Who gets what
# ═══════════════════════════════════════════════════════════════

REAL_CLIENT_CONTACTS = {
    "IT_ALL": ["jpetrokus@visualcomfort.com"],
    "D365_FO": ["mforsyth@visualcomfort.com"],
    "IT_Support_Savannah": ["vyates@visualcomfort.com"],
    "QA_QC": ["knwobu@visualcomfort.com"],
    "JDE_EDI": ["bbrewer@visualcomfort.com", "hchavarria@visualcomfort.com"],
    "Enterprise_Integration": ["apfister@visualcomfort.com"],
    "Salesforce": ["mdesai@visualcomfort.com"],
    "PartnerInsight": ["rgliane@visualcomfort.com"],
    "Web_B2B": ["rgliane@visualcomfort.com"],
}

def _client(group):
    """Return test or real client contacts depending on mode."""
    if TEST_MODE:
        return [TEST_EMAIL]
    return REAL_CLIENT_CONTACTS.get(group, [])

TIMESHEET_CONFIG = {
    "IT_ALL": {
        "filename": "Timesheet_VCC_IT_ALL",
        "client_contacts": _client("IT_ALL"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "IT_HelpDesk", "name": "Senthilnathan"},
            {"project": "IT_HelpDesk", "name": "Vivekanandan"},
            {"project": "IT_Infra_8X8", "name": "Aldrin"},
            {"project": "IT_Infra Onsite", "name": "Marcos"},
            {"project": "IT Infra PMO", "name": "Janani"},
            {"project": "IT Infra PMO", "name": "Janna"},
            {"project": "IT_infra_Security", "name": "Karthikeyan"},
            {"project": "IT_infra_Security", "name": "Mohammad"},
            {"project": "Projects_financial_services", "name": "Javal"},
        ],
    },
    "D365_FO": {
        "filename": "Timesheet_VCC_D365_FO",
        "client_contacts": _client("D365_FO"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "D365_FO", "name": "Manishkumar"},
            {"project": "D365_FO", "name": "Bradley"},
            {"project": "D365_FO", "name": "Kishore"},
            {"project": "D365_FO", "name": "Krupa"},
            {"project": "D365_FO", "name": "Andrea"},
            {"project": "D365_FO", "name": "Sankar"},
            {"project": "D365_FO", "name": "Anant"},
            {"project": "D365_FO", "name": "Shahul"},
            {"project": "D365_FO", "name": "Boyanarasimha"},
            {"project": "D365_FO", "name": "Srinivasan"},
            {"project": "D365_FO", "name": "Aravindh"},
            {"project": "D365_FO", "name": "Abhinandhan"},
            {"project": "D365_FO", "name": "Meenalochini"},
            {"project": "D365_FO", "name": "Keerthivasan"},
        ],
    },
    "IT_Support_Savannah": {
        "filename": "Timesheet_VCC_IT_Support_Savannah",
        "client_contacts": _client("IT_Support_Savannah"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "IT Savannah Support", "name": "Dhiraj"},
            {"project": "IT Savannah Support", "name": "Akhila"},
        ],
    },
    "QA_QC": {
        "filename": "Timesheet_VCC_QA_QC",
        "client_contacts": _client("QA_QC"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "QA QC", "name": "Bindu"},
            {"project": "QA QC", "name": "Saritha"},
            {"project": "QA QC", "name": "Ganesh"},
            {"project": "QA QC", "name": "Manjari"},
            {"project": "QA QC", "name": "Mohammed"},
            {"project": "Salesforce", "name": "Nishandhini"},
            {"project": "Web B2B", "name": "Aruldoss"},
            {"project": "D365_FO", "name": "Meenalochini"},
            {"project": "D365_FO", "name": "Shahul"},
            {"project": "PartnerInsight", "name": "Muthu"},
            {"project": "QA QC", "name": "Madhavi"},
            {"project": "QA QC", "name": "Bhavesh"},
            {"project": "QA Qc- Ecomm", "name": "Ashwath"},
        ],
    },
    "JDE_EDI": {
        "filename": "Timesheet_VCC_JDE_EDI",
        "client_contacts": _client("JDE_EDI"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "JDE & EDI", "name": "Arul"},
            {"project": "JDE & EDI", "name": "Thomas"},
            {"project": "JDE & EDI", "name": "Bholeshankar"},
            {"project": "JDE & EDI", "name": "Nitinkumar"},
            {"project": "JDE & EDI", "name": "Chandan"},
            {"project": "JDE & EDI", "name": "Irfan"},
        ],
    },
    "Enterprise_Integration": {
        "filename": "Timesheet_VCC_Enterprise_Integration",
        "client_contacts": _client("Enterprise_Integration"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "Enterprise_Integration", "name": "Pothiraja"},
            {"project": "Enterprise_Integration", "name": "Rafighafoor"},
        ],
    },
    "Salesforce": {
        "filename": "Timesheet_VCC_Salesforce",
        "client_contacts": _client("Salesforce"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "Salesforce", "name": "Divya"},
            {"project": "Salesforce", "name": "Nishandhini"},
            {"project": "Salesforce", "name": "Naveenkumar"},
            {"project": "Salesforce", "name": "Karthiga"},
            {"project": "Salesforce", "name": "Sandhirasegaran"},
            {"project": "Salesforce", "name": "Hari"},
            {"project": "Salesforce", "name": "Swaminathan"},
            {"project": "Salesforce", "name": "Mohammed"},
        ],
    },
    "PartnerInsight": {
        "filename": "Timesheet_VCC_PartnerInsight",
        "client_contacts": _client("PartnerInsight"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "PartnerInsight", "name": "Muthu", "role": "Test Engineer"},
            {"project": "PartnerInsight", "name": "Sasikumar", "role": "Full Stack Developer"},
            {"project": "PartnerInsight", "name": "Mahesh", "role": "Technical Lead"},
            {"project": "PartnerInsight", "name": "Balaji", "role": "Full Stack Developer"},
            {"project": "PartnerInsight", "name": "Kiran", "role": "Boomi Developer"},
            {"project": "PartnerInsight", "name": "Nageshwar", "role": "DB Developer/PM"},
            {"project": "PartnerInsight", "name": "Gayathri", "role": "UI/UX Web Designer"},
            {"project": "PartnerInsight", "name": "Humera", "role": "Onsite Developer"},
            {"project": "PartnerInsight", "name": "Zamir", "role": "BA"},
        ],
    },
    "Web_B2B": {
        "filename": "Timesheet_VCC_Web_B2B",
        "client_contacts": _client("Web_B2B"),
        "cc": [NICK_EMAIL, AR_EMAIL, ANAND_EMAIL, KISHAN_EMAIL],
        "resources": [
            {"project": "Web B2B", "name": "Aruldoss", "role": "QA"},
            {"project": "Web B2B", "name": "Jagadesh", "role": "Test Engineer"},
            {"project": "Web B2B", "name": "Sathishraj", "role": "Full Stack Developer"},
            {"project": "Web B2B", "name": "Vimal", "role": "Technical Lead"},
            {"project": "Web B2B", "name": "Anand S", "role": "Tiger Ops Support, PM"},
        ],
    },
}

# All resources flat list (for completion check)
ALL_RESOURCES = []
for group in TIMESHEET_CONFIG.values():
    for r in group["resources"]:
        ALL_RESOURCES.append(r["name"])


# ═══════════════════════════════════════════════════════════════
# GRAPH API HELPERS
# ═══════════════════════════════════════════════════════════════

_token_cache = {"token": None, "expires": 0}

def get_token():
    if _token_cache["token"] and time.time() < _token_cache["expires"]:
        return _token_cache["token"]
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = urllib.parse.urlencode({
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }).encode()
    req = urllib.request.Request(url, data=data, method="POST")
    with urllib.request.urlopen(req, context=ssl_ctx) as resp:
        d = json.loads(resp.read())
    _token_cache["token"] = d["access_token"]
    _token_cache["expires"] = time.time() + d["expires_in"] - 60
    return d["access_token"]


def graph_post(path, body):
    token = get_token()
    url = f"https://graph.microsoft.com/v1.0{path}" if not path.startswith("http") else path
    data = json.dumps(body).encode()
    req = urllib.request.Request(url, data=data, method="POST")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, context=ssl_ctx, timeout=30) as resp:
            raw = resp.read()
            if raw:
                return json.loads(raw)
            return {"status": "ok", "code": resp.status}
    except urllib.error.HTTPError as e:
        err_body = e.read().decode()[:500]
        raise Exception(f"Graph POST {path} failed ({e.code}): {err_body}")


def graph_patch(path, body):
    token = get_token()
    url = f"https://graph.microsoft.com/v1.0{path}" if not path.startswith("http") else path
    data = json.dumps(body).encode()
    req = urllib.request.Request(url, data=data, method="PATCH")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, context=ssl_ctx, timeout=30) as resp:
            raw = resp.read()
            if raw:
                return json.loads(raw)
            return {"status": "ok", "code": resp.status}
    except urllib.error.HTTPError as e:
        err_body = e.read().decode()[:500]
        raise Exception(f"Graph PATCH {path} failed ({e.code}): {err_body}")


def graph_get(path):
    token = get_token()
    url = f"https://graph.microsoft.com/v1.0{path}" if not path.startswith("http") else path
    # Encode spaces and special chars in query string for strict URL parsers
    if "?" in url:
        base, qs = url.split("?", 1)
        url = base + "?" + urllib.parse.quote(qs, safe="=&$,/':")
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, context=ssl_ctx, timeout=30) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        err_body = e.read().decode()[:500]
        raise Exception(f"Graph GET {path} failed ({e.code}): {err_body}")


# ═══════════════════════════════════════════════════════════════
# STEP 1: COMPLETION CHECK & TEAMS PING
# ═══════════════════════════════════════════════════════════════

def get_current_period():
    """Get the current 2-week period label."""
    today = datetime.now()
    # Periods: 1st-15th, 16th-end of month
    if today.day <= 15:
        start = today.replace(day=1)
        end = today.replace(day=15)
    else:
        start = today.replace(day=16)
        # Last day of month
        if today.month == 12:
            end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    return f"{start.strftime('%d%b')}To-{end.strftime('%d%b')}_{today.year}"


def find_user_by_name(name):
    """Find a user's email/ID in Azure AD by display name (with ConsistencyLevel header)."""
    token = get_token()
    search = urllib.parse.quote(name)

    # Try $search with ConsistencyLevel: eventual (required by Graph API)
    url = f"https://graph.microsoft.com/v1.0/users?$search=%22displayName:{search}%22&$select=id,displayName,mail&$count=true&$top=5"
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("ConsistencyLevel", "eventual")
    try:
        with urllib.request.urlopen(req, context=ssl_ctx, timeout=30) as resp:
            result = json.loads(resp.read())
        users = result.get("value", [])
        if users:
            return users[0]
    except:
        pass

    # Fallback: try startswith filter on displayName
    filter_str = urllib.parse.quote(f"startswith(displayName,'{name}')")
    try:
        result = graph_get(f"/users?$filter={filter_str}&$select=id,displayName,mail&$top=5")
        users = result.get("value", [])
        if users:
            return users[0]
    except:
        pass

    return None


def send_teams_message(user_email, message):
    """Send a Teams chat message to a user."""
    try:
        # Create 1:1 chat
        chat = graph_post("/chats", {
            "chatType": "oneOnOne",
            "members": [
                {
                    "@odata.type": "#microsoft.graph.aadUserConversationMember",
                    "roles": ["owner"],
                    "user@odata.bind": f"https://graph.microsoft.com/v1.0/users/{user_email}",
                },
                {
                    "@odata.type": "#microsoft.graph.aadUserConversationMember",
                    "roles": ["owner"],
                    "user@odata.bind": f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}",
                },
            ],
        })
        # Send message
        graph_post(f"/chats/{chat['id']}/messages", {
            "body": {"contentType": "html", "content": message},
        })
        return True
    except Exception as e:
        print(f"    Failed to message {user_email}: {e}")
        return False


def check_completion(period_label, submitted_names):
    """
    Check who has submitted and ping those who haven't.
    submitted_names: list of names who HAVE submitted.
    """
    missing = [name for name in ALL_RESOURCES if name not in submitted_names]

    if not missing:
        print(f"\n  ALL {len(ALL_RESOURCES)} resources have submitted for {period_label}")
        print("  Period can be FROZEN.")
        return True

    print(f"\n  {len(ALL_RESOURCES) - len(missing)}/{len(ALL_RESOURCES)} submitted")
    print(f"  MISSING ({len(missing)}):")
    for name in missing:
        print(f"    - {name}")

    # Ping missing via Teams
    confirm = input("\n  Send Teams reminder to missing members? [y/N]: ").strip().lower()
    if confirm == "y":
        message = (
            f"<b>Timesheet Reminder</b><br><br>"
            f"Hi, your timesheet for period <b>{period_label}</b> has not been submitted yet. "
            f"Please submit it ASAP so we can proceed with consolidation.<br><br>"
            f"Thank you,<br>Gayathri"
        )
        for name in missing:
            user = find_user_by_name(name)
            if user and user.get("mail"):
                print(f"    Pinging {name} ({user['mail']})...", end=" ")
                if send_teams_message(user["mail"], message):
                    print("SENT")
                else:
                    print("FAILED")
            else:
                print(f"    Could not find {name} in Azure AD")

    return False


# ═══════════════════════════════════════════════════════════════
# STEPS 2-5: GENERATE CONSOLIDATED EXCEL
# ═══════════════════════════════════════════════════════════════

def parse_period_dates(period_label):
    """Parse period label like '13AprTo-26Apr_2026' into start/end dates and working days."""
    import re as _re
    m = _re.match(r'(\d+)([A-Za-z]+)To-(\d+)([A-Za-z]+)_(\d{4})', period_label)
    if not m:
        # Fallback: 10 generic working days
        return None, None, []
    start_day, start_mon, end_day, end_mon, year = m.groups()
    month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                 "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    start = datetime(int(year), month_map[start_mon.lower()], int(start_day))
    end = datetime(int(year), month_map[end_mon.lower()], int(end_day))
    # Collect working days (Mon-Fri)
    working_days = []
    d = start
    while d <= end:
        if d.weekday() < 5:  # Mon=0 .. Fri=4
            working_days.append(d)
        d += timedelta(days=1)
    return start, end, working_days


def create_project_summary_sheet(wb, group_key, resources_with_hours, period_label):
    """Create the Summary sheet (first tab) in a project workbook."""
    ws = wb.create_sheet(title="Summary", index=0)

    has_role = any(r.get("role") for r in resources_with_hours)

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    bold = Font(bold=True, size=11)

    # Title row
    ws.cell(row=1, column=1, value=f"Timesheet — {group_key.replace('_', ' ')}").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Period: {period_label}").font = Font(italic=True, size=10, color="666666")

    # Headers (row 4)
    hdr_row = 4
    if has_role:
        headers = ["Sl.No", "Project", "Resource", "Role", "Hours"]
    else:
        headers = ["Sl.No", "Project", "Resource", "Hours"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    total_hours = 0
    for i, r in enumerate(resources_with_hours, 1):
        row = hdr_row + i
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=r["project"]).border = thin_border
        ws.cell(row=row, column=3, value=r["name"]).border = thin_border
        if has_role:
            ws.cell(row=row, column=4, value=r.get("role", "")).border = thin_border
            hours_cell = ws.cell(row=row, column=5, value=r.get("hours", 0))
        else:
            hours_cell = ws.cell(row=row, column=4, value=r.get("hours", 0))
        hours_cell.border = thin_border
        hours_cell.alignment = Alignment(horizontal="right")
        total_hours += r.get("hours", 0)

    # Total row
    total_row = hdr_row + len(resources_with_hours) + 1
    hours_col = 5 if has_role else 4
    ws.cell(row=total_row, column=2, value="TOTAL").font = bold
    ws.cell(row=total_row, column=2).border = thin_border
    ws.cell(row=total_row, column=3, value=f"{len(resources_with_hours)} resources").font = bold
    ws.cell(row=total_row, column=3).border = thin_border
    total_cell = ws.cell(row=total_row, column=hours_col, value=total_hours)
    total_cell.font = bold
    total_cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    if has_role:
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 12
    else:
        ws.column_dimensions["D"].width = 12

    return total_hours


def create_person_sheet(wb, resource, period_label, working_days):
    """Create an individual sheet for a person with daily breakdown."""
    name = resource["name"]
    project = resource["project"]
    hours = resource.get("hours", 0)
    role = resource.get("role", "")

    # Sheet name: first 31 chars (Excel limit)
    sheet_title = name[:31]
    # Avoid duplicate sheet names
    if sheet_title in wb.sheetnames:
        sheet_title = f"{name[:27]}_{project[:3]}"[:31]
    ws = wb.create_sheet(title=sheet_title)

    # Styles
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    bold = Font(bold=True, size=10)
    light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Info header
    ws.cell(row=1, column=1, value="Vendor Name:").font = bold
    ws.cell(row=1, column=2, value="D4 Insight Inc.")
    ws.cell(row=2, column=1, value="Contractor Name:").font = bold
    ws.cell(row=2, column=2, value=name)
    ws.cell(row=3, column=1, value="Project:").font = bold
    ws.cell(row=3, column=2, value=project)
    if role:
        ws.cell(row=4, column=1, value="Role:").font = bold
        ws.cell(row=4, column=2, value=role)
    ws.cell(row=5, column=1, value="Period:").font = bold
    ws.cell(row=5, column=2, value=period_label)

    # Table headers (row 7)
    hdr_row = 7
    headers = ["Date", "Vendor Name", "Contractor Name", "Project Name", "Work Item", "Description", "Hours"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Distribute hours across working days
    num_days = len(working_days) if working_days else 10
    if num_days > 0 and hours > 0:
        daily_hours = hours / num_days
        # Round to nearest 0.5, ensure total matches
        daily_list = [round(daily_hours * 2) / 2] * num_days
        # Adjust last day to match exact total
        daily_list[-1] = hours - sum(daily_list[:-1])
    else:
        daily_list = [0] * num_days

    # Data rows
    for i, (day, dh) in enumerate(zip(working_days if working_days else range(num_days), daily_list)):
        row = hdr_row + 1 + i
        if isinstance(day, datetime):
            date_str = day.strftime("%d-%b-%Y")
        else:
            date_str = f"Day {day + 1}"

        fill = light_fill if i % 2 == 0 else None
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
            if fill:
                ws.cell(row=row, column=col).fill = fill

        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value="D4 Insight Inc.")
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=project)
        ws.cell(row=row, column=5, value=project)  # Work Item = project name
        ws.cell(row=row, column=6, value=f"{project} - Daily work")
        hrs_cell = ws.cell(row=row, column=7, value=round(dh, 2))
        hrs_cell.alignment = Alignment(horizontal="right")

    # Total row
    total_row = hdr_row + 1 + max(len(working_days), num_days)
    ws.cell(row=total_row, column=6, value="TOTAL").font = bold
    ws.cell(row=total_row, column=6).border = thin_border
    ws.cell(row=total_row, column=7, value=hours).font = bold
    ws.cell(row=total_row, column=7).border = thin_border
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 10


def create_master_summary_sheet(wb, project_summaries):
    """Create the overall Summary tab for the master consolidated file."""
    ws = wb.create_sheet(title="Summary", index=0)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Project Group", "Total Hours", "No. of Resources"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    grand_total_hours = 0
    grand_total_resources = 0
    for i, s in enumerate(project_summaries, 2):
        ws.cell(row=i, column=1, value=s["project"]).border = thin_border
        ws.cell(row=i, column=2, value=s["hours"]).border = thin_border
        ws.cell(row=i, column=3, value=s["resources"]).border = thin_border
        grand_total_hours += s["hours"]
        grand_total_resources += s["resources"]

    total_row = len(project_summaries) + 2
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=2, value=grand_total_hours).font = Font(bold=True)
    ws.cell(row=total_row, column=2).border = thin_border
    ws.cell(row=total_row, column=3, value=grand_total_resources).font = Font(bold=True)
    ws.cell(row=total_row, column=3).border = thin_border

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18


def generate_excel(period_label, timesheet_data=None):
    """
    Generate Excel files per project group, each with:
    - Summary sheet (consolidated table)
    - Individual sub-sheets per person (daily breakdown)
    Plus a master consolidated file with overall summary.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    period_dir = os.path.join(OUTPUT_DIR, period_label)
    os.makedirs(period_dir, exist_ok=True)

    # Parse period dates for daily breakdown
    _, _, working_days = parse_period_dates(period_label)

    generated_files = []
    all_summaries = []
    validation_issues = []

    for group_key, config in TIMESHEET_CONFIG.items():
        # Get hours data
        resources = []
        for r in config["resources"]:
            hours = 0
            if timesheet_data and group_key in timesheet_data:
                match = next((t for t in timesheet_data[group_key] if t["name"].lower() == r["name"].lower()), None)
                if match:
                    hours = match.get("hours", 0)
            resources.append({**r, "hours": hours})

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1) Summary sheet (first tab)
        total_hours = create_project_summary_sheet(wb, group_key, resources, period_label)

        # 2) Individual person sheets
        for r in resources:
            create_person_sheet(wb, r, period_label, working_days)

        # Save
        filename = f"{config['filename']}_{period_label}.xlsx"
        filepath = os.path.join(period_dir, filename)
        wb.save(filepath)
        generated_files.append(filepath)

        all_summaries.append({
            "project": group_key,
            "hours": total_hours,
            "resources": len(resources),
        })

        # Validation
        for r in resources:
            if r["hours"] == 0:
                validation_issues.append(f"  [MISSING] {r['name']} in {group_key}: 0 hours")
            elif r["hours"] > 96:
                validation_issues.append(f"  [HIGH] {r['name']} in {group_key}: {r['hours']} hours (>96)")
            elif r["hours"] < 40:
                validation_issues.append(f"  [LOW] {r['name']} in {group_key}: {r['hours']} hours (<40)")

        print(f"  Generated: {filename} ({len(resources)} resources, {total_hours}h)")

    # Master consolidated file: Summary + per-group summary sheets
    master_wb = openpyxl.Workbook()
    master_wb.remove(master_wb.active)
    create_master_summary_sheet(master_wb, all_summaries)

    for group_key, config in TIMESHEET_CONFIG.items():
        resources = []
        for r in config["resources"]:
            hours = 0
            if timesheet_data and group_key in timesheet_data:
                match = next((t for t in timesheet_data[group_key] if t["name"].lower() == r["name"].lower()), None)
                if match:
                    hours = match.get("hours", 0)
            resources.append({**r, "hours": hours})
        create_project_summary_sheet(master_wb, group_key, resources, period_label)

    master_path = os.path.join(period_dir, f"Consolidated_VCC_ALL_{period_label}.xlsx")
    master_wb.save(master_path)
    generated_files.append(master_path)
    print(f"\n  Master file: {master_path}")

    # Validation report
    if validation_issues:
        print(f"\n  VALIDATION ISSUES ({len(validation_issues)}):")
        for issue in validation_issues[:20]:
            print(issue)

    # Check for duplicates
    name_groups = {}
    for group_key, config in TIMESHEET_CONFIG.items():
        for r in config["resources"]:
            name_groups.setdefault(r["name"], []).append(group_key)
    duplicates = {name: groups for name, groups in name_groups.items() if len(groups) > 1}
    if duplicates:
        print(f"\n  DUPLICATE RESOURCES (across sheets):")
        for name, groups in duplicates.items():
            print(f"    {name}: appears in {', '.join(groups)}")

    return generated_files, period_dir


# ═══════════════════════════════════════════════════════════════
# STEP 6: SEND TO ANAND FOR REVIEW
# ═══════════════════════════════════════════════════════════════

def send_email(from_email, to_emails, cc_emails, subject, html_body, attachments=None, reply_to_message_id=None):
    """Send email via Graph API with optional attachments. If reply_to_message_id is provided, replies in-thread."""
    import base64

    # SAFETY GUARD: Only allow internal @d4insight.com recipients from sysadmin
    all_recipients = list(to_emails or []) + list(cc_emails or [])
    blocked = [e for e in all_recipients if not e.strip().lower().endswith("@d4insight.com")]
    if blocked:
        raise Exception(
            f"BLOCKED: Cannot send to external emails via sysadmin: {blocked}. "
            f"Only @d4insight.com addresses are permitted."
        )

    # If replying to an existing message, use createReply (draft) -> patch body -> send
    # This preserves threading headers (In-Reply-To, References, conversationId)
    if reply_to_message_id:
        # Step 1: Create a reply draft (preserves conversation threading)
        draft = graph_post(f"/users/{from_email}/messages/{reply_to_message_id}/createReply", {})
        draft_id = draft.get("id")
        if not draft_id:
            raise Exception("createReply did not return a draft ID")
        # Step 2: Update the draft body (and optionally recipients)
        patch_body = {"body": {"contentType": "HTML", "content": html_body}}
        if to_emails:
            patch_body["toRecipients"] = [{"emailAddress": {"address": e}} for e in to_emails]
        if cc_emails:
            patch_body["ccRecipients"] = [{"emailAddress": {"address": e}} for e in cc_emails]
        graph_patch(f"/users/{from_email}/messages/{draft_id}", patch_body)
        # Step 3: Send the draft
        graph_post(f"/users/{from_email}/messages/{draft_id}/send", {})
        return

    message = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html_body},
        "toRecipients": [{"emailAddress": {"address": e}} for e in to_emails],
    }
    if cc_emails:
        message["ccRecipients"] = [{"emailAddress": {"address": e}} for e in cc_emails]

    if attachments:
        message["attachments"] = []
        for filepath in attachments:
            with open(filepath, "rb") as f:
                content = base64.b64encode(f.read()).decode()
            message["attachments"].append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": os.path.basename(filepath),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentBytes": content,
            })

    graph_post(f"/users/{from_email}/sendMail", {"message": message})


GAYATHRI_EMAIL = "gayathri.m@d4insight.com"


def send_to_anand(period_dir, period_label):
    """Send all generated files to Anand AND Gayathri for review and approval."""
    files = [os.path.join(period_dir, f) for f in os.listdir(period_dir) if f.endswith(".xlsx")]

    html = f"""
    <p>Hi,</p>
    <p>Please find attached the consolidated timesheets for period <b>{period_label}</b> for your review and approval.</p>
    <p>Kindly review and let me know if any changes are needed.</p>
    <br>
    <p>Thank you and regards,<br>Qoppy — D4 Insight Automation</p>
    """

    to_list = [ANAND_EMAIL, GAYATHRI_EMAIL]
    # Deduplicate (in TEST_MODE both might resolve to same email)
    to_list = list(dict.fromkeys(to_list))

    send_email(
        from_email=SENDER_EMAIL,
        to_emails=to_list,
        cc_emails=[KISHAN_EMAIL],
        subject=f"Timesheets for approval - {period_label}",
        html_body=html,
        attachments=files,
    )
    print(f"\n  Sent {len(files)} files to {', '.join(to_list)}")


# ═══════════════════════════════════════════════════════════════
# STEP 7: SEND FINAL APPROVED TIMESHEETS TO GAYATHRI
# ═══════════════════════════════════════════════════════════════

def send_final_to_gayathri(period_dir, period_label):
    """Send all approved/final Excel files to Gayathri for processing."""
    files = [os.path.join(period_dir, f) for f in os.listdir(period_dir) if f.endswith(".xlsx")]

    if not files:
        print("  No Excel files found to send.")
        return

    html = f"""
    <p>Hi Gayathri,</p>
    <p>Please find attached the <b>approved and finalized</b> timesheets for period <b>{period_label}</b>.</p>
    <p>These have been reviewed and approved. Please proceed with processing.</p>
    <br>
    <p>Regards,<br>Qoppy — D4 Insight Automation</p>
    """

    send_email(
        from_email=SENDER_EMAIL,
        to_emails=[GAYATHRI_EMAIL],
        cc_emails=[ANAND_EMAIL, KISHAN_EMAIL],
        subject=f"APPROVED Timesheets - {period_label} (Final)",
        html_body=html,
        attachments=files,
    )
    print(f"\n  Sent {len(files)} final files to {GAYATHRI_EMAIL}")


# ═══════════════════════════════════════════════════════════════
# LEAVE TRACKER
# ═══════════════════════════════════════════════════════════════

def _load_leave_data():
    """Load leave records from JSON file."""
    os.makedirs(LEAVE_DATA_DIR, exist_ok=True)
    data_file = os.path.join(LEAVE_DATA_DIR, "leave_records.json")
    if os.path.exists(data_file):
        with open(data_file) as f:
            return json.load(f)
    return {"records": [], "last_checked_id": None}


def _save_leave_data(data):
    """Save leave records to JSON file."""
    os.makedirs(LEAVE_DATA_DIR, exist_ok=True)
    data_file = os.path.join(LEAVE_DATA_DIR, "leave_records.json")
    with open(data_file, "w") as f:
        json.dump(data, f, indent=2)


def _match_resource_by_email(email):
    """Match an email address to a resource name."""
    email_lower = email.lower()
    # Try to find user in Azure AD and match to our resource list
    for group_key, config in TIMESHEET_CONFIG.items():
        for r in config["resources"]:
            name_lower = r["name"].lower()
            # Check if resource name is part of the email (e.g., senthilnathan in senthilnathan.r@d4insight.com)
            if name_lower in email_lower:
                return r["name"], group_key
    return None, None


def _parse_leave_dates_from_subject(subject, body_text=""):
    """
    Parse leave dates from email subject/body.
    Handles formats like:
    - "Leave on 23rd April"
    - "Leave 23-25 Apr 2026"
    - "On leave April 23"
    - "WFH 23 Apr" (not leave)
    - "Sick leave 23/04/2026"
    """
    import re as _re
    text = f"{subject} {body_text}".lower()

    # Skip if it's just WFH
    if "wfh" in text and "leave" not in text:
        return []

    dates = []
    today = datetime.now()
    year = today.year

    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
        "january": 1, "february": 2, "march": 3, "april": 4,
        "june": 6, "july": 7, "august": 8, "september": 9,
        "october": 10, "november": 11, "december": 12,
    }

    # Pattern: "23rd April" or "23 Apr" or "April 23"
    for m in _re.finditer(r'(\d{1,2})(?:st|nd|rd|th)?\s+(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)', text):
        day = int(m.group(1))
        mon = month_map.get(m.group(2))
        if mon and 1 <= day <= 31:
            try:
                dates.append(datetime(year, mon, day))
            except ValueError:
                pass

    # Pattern: "April 23" (month first)
    for m in _re.finditer(r'(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s+(\d{1,2})(?:st|nd|rd|th)?', text):
        mon = month_map.get(m.group(1))
        day = int(m.group(2))
        if mon and 1 <= day <= 31:
            try:
                d = datetime(year, mon, day)
                if d not in dates:
                    dates.append(d)
            except ValueError:
                pass

    # Pattern: date range "23-25 Apr" or "23 to 25 Apr"
    for m in _re.finditer(r'(\d{1,2})(?:st|nd|rd|th)?\s*(?:-|to)\s*(\d{1,2})(?:st|nd|rd|th)?\s+(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)', text):
        start_day = int(m.group(1))
        end_day = int(m.group(2))
        mon = month_map.get(m.group(3))
        if mon:
            for d in range(start_day, end_day + 1):
                try:
                    dt = datetime(year, mon, d)
                    if dt not in dates:
                        dates.append(dt)
                except ValueError:
                    pass

    # Pattern: dd/mm/yyyy
    for m in _re.finditer(r'(\d{1,2})/(\d{1,2})/(\d{4})', text):
        day, mon, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            dt = datetime(yr, mon, day)
            if dt not in dates:
                dates.append(dt)
        except ValueError:
            pass

    # If no specific dates found but "leave" or "sick" mentioned, assume today
    if not dates and ("leave" in text or "sick" in text):
        dates.append(today)

    return sorted(dates)


def scan_leave_emails(period_label=None):
    """
    Scan leave tracker mailbox for new leave notifications.
    Returns list of parsed leave records.
    """
    data = _load_leave_data()
    last_id = data.get("last_checked_id")

    print(f"\n  Scanning {LEAVE_TRACKER_EMAIL} for leave emails...\n")

    # Get recent emails (last 30 days)
    thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%dT00:00:00Z")
    url = f"/users/{LEAVE_TRACKER_EMAIL}/messages?$filter=receivedDateTime ge {thirty_days_ago}&$top=100&$select=id,from,subject,receivedDateTime,body,ccRecipients"

    try:
        result = graph_get(url)
    except Exception as e:
        print(f"  Error reading mailbox: {e}")
        return []

    messages = result.get("value", [])
    print(f"  Found {len(messages)} email(s) in last 30 days")

    # Track existing record IDs to avoid duplicates
    existing_ids = {r.get("email_id") for r in data["records"]}

    new_leaves = []
    for msg in messages:
        msg_id = msg.get("id", "")
        if msg_id in existing_ids:
            continue

        from_addr = msg.get("from", {}).get("emailAddress", {}).get("address", "")
        from_name = msg.get("from", {}).get("emailAddress", {}).get("name", "")
        subject = msg.get("subject", "")
        received = msg.get("receivedDateTime", "")
        body_raw = msg.get("body", {}).get("content", "")

        import re as _re
        body_text = _re.sub(r'<[^>]+>', '', body_raw).strip()[:1000]

        # Skip non-leave emails (SOW approvals, timesheets, meeting notes, etc.)
        subject_lower = subject.lower()
        skip_keywords = ['sow approval', 'sow revised', 'statement of work', 'timesheet', 'meeting notes', 'invoice', 'docusign']
        if any(kw in subject_lower for kw in skip_keywords):
            continue
        # Only process if subject or body mentions leave/vacation/time off/absence
        leave_keywords = ['leave', 'vacation', 'time off', 'time-off', 'pto', 'absent', 'absence', 'sick', 'out of office', 'ooo', 'wfh']
        if not any(kw in subject_lower or kw in body_text.lower()[:500] for kw in leave_keywords):
            continue

        # Check if leave tracker was CC'd (or if mail was sent TO leave tracker)
        cc_addrs = [c.get("emailAddress", {}).get("address", "").lower()
                     for c in msg.get("ccRecipients", [])]

        # Parse who is on leave
        resource_name, group = _match_resource_by_email(from_addr)
        if not resource_name:
            # Try matching from display name
            for gk, cfg in TIMESHEET_CONFIG.items():
                for r in cfg["resources"]:
                    if r["name"].lower() in from_name.lower() or r["name"].lower() in subject.lower():
                        resource_name = r["name"]
                        group = gk
                        break
                if resource_name:
                    break

        # Parse dates
        leave_dates = _parse_leave_dates_from_subject(subject, body_text)

        if resource_name and leave_dates:
            record = {
                "email_id": msg_id,
                "resource": resource_name,
                "group": group,
                "email": from_addr,
                "subject": subject,
                "received": received,
                "leave_dates": [d.strftime("%Y-%m-%d") for d in leave_dates],
                "num_days": len(leave_dates),
                "status": "active",
            }
            new_leaves.append(record)
            data["records"].append(record)
            print(f"  NEW: {resource_name} — {len(leave_dates)} day(s): {', '.join(d.strftime('%d-%b') for d in leave_dates)}")

            # Auto-sync to Supabase dashboard
            sync_leave_to_supabase(
                resource_name,
                [d.strftime("%Y-%m-%d") for d in leave_dates],
                reason=subject,
            )

            # Auto-reply disabled — do not send any acknowledgment emails
            date_str = ', '.join(d.strftime('%d-%b-%Y') for d in leave_dates)
            print(f"    Recorded (no auto-reply): {date_str}")

        elif "leave" in subject.lower() or "leave" in body_text.lower():
            print(f"  UNMATCHED: {from_addr} — \"{subject[:60]}\" (could not match resource/dates)")

    data["last_checked_id"] = messages[0]["id"] if messages else last_id
    _save_leave_data(data)

    if not new_leaves:
        print("  No new leave entries found.")

    return new_leaves


def get_leaves_for_period(period_label):
    """Get all leave records that fall within a specific period."""
    start, end, _ = parse_period_dates(period_label)
    if not start or not end:
        return []

    data = _load_leave_data()
    period_leaves = []

    for record in data["records"]:
        if record.get("status") != "active":
            continue
        for date_str in record.get("leave_dates", []):
            try:
                d = datetime.strptime(date_str, "%Y-%m-%d")
                if start <= d <= end:
                    period_leaves.append(record)
                    break
            except ValueError:
                pass

    return period_leaves


def get_leave_summary(period_label=None):
    """Get leave count per resource, optionally filtered by period."""
    data = _load_leave_data()
    summary = {}

    for record in data["records"]:
        if record.get("status") != "active":
            continue
        name = record["resource"]
        dates = record.get("leave_dates", [])

        if period_label:
            start, end, _ = parse_period_dates(period_label)
            if start and end:
                dates = [d for d in dates if start <= datetime.strptime(d, "%Y-%m-%d") <= end]

        if dates:
            if name not in summary:
                summary[name] = {"resource": name, "group": record.get("group", ""), "dates": [], "total_days": 0}
            summary[name]["dates"].extend(dates)
            summary[name]["total_days"] = len(set(summary[name]["dates"]))

    return summary


def adjust_timesheet_for_leaves(period_label, timesheet_data):
    """
    Adjust timesheet hours based on leave records.
    Deducts 8 hours per leave day from the resource's total.
    Returns updated timesheet_data and list of adjustments made.
    """
    leaves = get_leaves_for_period(period_label)
    if not leaves:
        return timesheet_data, []

    start, end, working_days = parse_period_dates(period_label)
    num_working_days = len(working_days)
    adjustments = []

    # Count leave days per resource within the period
    leave_days_per_resource = {}
    for record in leaves:
        name = record["resource"]
        for date_str in record.get("leave_dates", []):
            d = datetime.strptime(date_str, "%Y-%m-%d")
            if start <= d <= end and d.weekday() < 5:  # Only working days
                leave_days_per_resource.setdefault(name, set()).add(date_str)

    # Adjust hours
    for name, leave_date_set in leave_days_per_resource.items():
        num_leave = len(leave_date_set)
        hours_to_deduct = num_leave * 8

        for group_key in timesheet_data:
            for r in timesheet_data[group_key]:
                if r["name"].lower() == name.lower():
                    old_hours = r.get("hours", 0)
                    new_hours = max(0, old_hours - hours_to_deduct)
                    r["hours"] = new_hours
                    adjustments.append({
                        "resource": name,
                        "group": group_key,
                        "leave_days": num_leave,
                        "old_hours": old_hours,
                        "new_hours": new_hours,
                        "dates": sorted(leave_date_set),
                    })
                    print(f"  Leave adjustment: {name} ({group_key}) {old_hours}h → {new_hours}h ({num_leave} day(s) leave)")

    return timesheet_data, adjustments


def check_excessive_leaves_and_notify(period_label=None):
    """
    Check if any resource has excessive leaves and auto-notify their manager.
    """
    summary = get_leave_summary(period_label)
    notifications_sent = []

    for name, info in summary.items():
        total = info["total_days"]
        group = info.get("group", "")

        if total >= LEAVE_THRESHOLD_CRITICAL:
            level = "CRITICAL"
        elif total >= LEAVE_THRESHOLD_WARNING:
            level = "WARNING"
        else:
            continue

        manager_email = RESOURCE_MANAGERS.get(name, DEFAULT_MANAGER)
        if TEST_MODE:
            manager_email = TEST_EMAIL

        period_text = f" for period {period_label}" if period_label else ""
        dates_str = ", ".join(sorted(set(info["dates"])))

        html = (
            f"<p>Hi,</p>"
            f"<p>This is an automatic notification regarding excessive leave{period_text}.</p>"
            f"<p><b>{name}</b> ({group}) has taken <b>{total} day(s)</b> of leave:</p>"
            f"<p>Dates: {dates_str}</p>"
        )
        if level == "CRITICAL":
            html += f"<p style='color:red'><b>⚠ CRITICAL:</b> {total} days exceeds the {LEAVE_THRESHOLD_CRITICAL}-day threshold. Immediate attention required.</p>"
        else:
            html += f"<p style='color:orange'><b>⚠ WARNING:</b> {total} days exceeds the {LEAVE_THRESHOLD_WARNING}-day threshold.</p>"

        html += (
            f"<p>Please review and take necessary action.</p>"
            f"<br><p>Regards,<br>Timesheet Automation System</p>"
        )

        cc_list = []
        if level == "CRITICAL":
            cc_list = [ANAND_EMAIL]

        print(f"  [{level}] {name}: {total} days leave → notifying {manager_email}")
        try:
            send_email(
                from_email=SENDER_EMAIL,
                to_emails=[manager_email],
                cc_emails=cc_list,
                subject=f"[{level}] Excessive Leave - {name} - {total} days{period_text}",
                html_body=html,
            )
            notifications_sent.append({"resource": name, "days": total, "level": level, "manager": manager_email})
        except Exception as e:
            print(f"    Failed to notify: {e}")

    return notifications_sent


def leave_status(period_label=None):
    """Display leave summary and status."""
    data = _load_leave_data()
    total_records = len(data["records"])
    active = [r for r in data["records"] if r.get("status") == "active"]

    print(f"\n  Leave Tracker Status")
    print(f"  {'='*50}")
    print(f"  Total records: {total_records}")
    print(f"  Active: {len(active)}")
    print(f"  Mailbox: {LEAVE_TRACKER_EMAIL}")

    if period_label:
        summary = get_leave_summary(period_label)
        if summary:
            print(f"\n  Leaves in period {period_label}:")
            print(f"  {'─'*50}")
            for name, info in sorted(summary.items()):
                flag = ""
                if info["total_days"] >= LEAVE_THRESHOLD_CRITICAL:
                    flag = " ⚠ CRITICAL"
                elif info["total_days"] >= LEAVE_THRESHOLD_WARNING:
                    flag = " ⚠ WARNING"
                dates = ", ".join(sorted(set(info["dates"])))
                print(f"    {name} ({info['group']}): {info['total_days']} day(s){flag}")
                print(f"      Dates: {dates}")
        else:
            print(f"\n  No leaves recorded for {period_label}")

    if active:
        print(f"\n  Recent leave records:")
        print(f"  {'─'*50}")
        for r in active[-10:]:
            dates = ", ".join(r.get("leave_dates", []))
            print(f"    {r['resource']} ({r.get('group','')}): {r['num_days']} day(s) — {dates}")
            print(f"      Subject: {r['subject'][:60]}")


def sync_leave_to_supabase(resource_name, leave_dates, leave_type="casual", reason=""):
    """
    Sync a leave record to Supabase so it appears in the dashboard.
    Maps resource name → Supabase user_id, then inserts into 'leaves' table
    and updates 'leave_balances'.
    """
    sb = get_supabase()
    if not sb:
        return False

    # Find user by name match (search all active users, not just employees)
    try:
        result = sb.table("users").select("id,name,email").eq("is_active", True).execute()
    except Exception as e:
        print(f"    Supabase error: {e}")
        return False

    user_id = None
    res_lower = resource_name.lower().replace(" ", "")
    for u in result.data:
        uname = u["name"].lower()
        # Match: "Senthilnathan" matches "Senthil Nathan Rajagopal"
        # Match: "Boyanarasimha" matches "Boya Narasimha Reddy"
        uname_nospace = uname.replace(" ", "")
        if res_lower in uname_nospace or uname_nospace.startswith(res_lower[:6]):
            user_id = u["id"]
            break
        # Also try matching first name
        first_name = uname.split()[0]
        if res_lower.startswith(first_name) or first_name.startswith(res_lower[:5]):
            user_id = u["id"]
            break

    if not user_id:
        print(f"    Could not find {resource_name} in Supabase users")
        return False

    # Get matched user name for the record
    matched_user_name = resource_name
    for u in result.data:
        if u["id"] == user_id:
            matched_user_name = u["name"]
            break

    # Calculate date range
    sorted_dates = sorted(leave_dates)
    start_date = sorted_dates[0]
    end_date = sorted_dates[-1]
    # Count working days
    days_count = 0
    for d_str in sorted_dates:
        d = datetime.strptime(d_str, "%Y-%m-%d")
        if d.weekday() < 5:
            days_count += 1

    if days_count == 0:
        return False

    # Insert into leaves table
    leave_record = {
        "user_id": user_id,
        "user_name": matched_user_name,
        "leave_type": leave_type,
        "start_date": start_date,
        "end_date": end_date,
        "days_count": days_count,
        "reason": reason or "Leave detected from email",
        "status": "approved",  # Auto-approved since it came from email notification
    }

    try:
        sb.table("leaves").insert(leave_record).execute()
        print(f"    Dashboard synced: {resource_name} ({user_id}) — {days_count} day(s)")

        # Update leave balance
        bal_id = f"{user_id}_{leave_type}"
        bal_result = sb.table("leave_balances").select("*").eq("id", bal_id).execute()
        if bal_result.data:
            old_used = bal_result.data[0].get("used_days", 0)
            sb.table("leave_balances").update({"used_days": old_used + days_count}).eq("id", bal_id).execute()

        return True
    except Exception as e:
        print(f"    Supabase sync error: {e}")
        return False


def sync_timesheet_to_supabase(period_label, timesheet_data):
    """
    Sync timesheet hours to Supabase so the dashboard reflects current data.
    """
    sb = get_supabase()
    if not sb:
        return

    # Get all users for mapping
    try:
        users_result = sb.table("users").select("id,name").eq("role", "employee").execute()
    except Exception as e:
        print(f"  Supabase sync error: {e}")
        return

    name_to_id = {}
    for u in users_result.data:
        name_to_id[u["name"].lower()] = u["id"]
        # Also map by first name
        first = u["name"].split()[0].lower()
        if first not in name_to_id:
            name_to_id[first] = u["id"]

    start, end, working_days = parse_period_dates(period_label)
    if not start:
        return

    period_key = f"{start.strftime('%Y-%m-%d')}_{end.strftime('%Y-%m-%d')}"
    synced = 0

    for group_key, resources in timesheet_data.items():
        for r in resources:
            name_lower = r["name"].lower()
            user_id = name_to_id.get(name_lower)
            if not user_id:
                continue

            hours = r.get("hours", 0)
            if hours <= 0:
                continue

            # Create timesheet entries for each working day
            daily_hours = hours / len(working_days) if working_days else 0

            for day in working_days:
                date_str = day.strftime("%Y-%m-%d")
                entry = {
                    "user_id": user_id,
                    "date": date_str,
                    "project": r.get("project", group_key),
                    "hours": round(daily_hours, 2),
                    "description": f"{r.get('project', group_key)} - Daily work",
                    "status": "submitted",
                }
                try:
                    sb.table("timesheets").upsert(entry, on_conflict="user_id,date,project").execute()
                    synced += 1
                except:
                    pass  # Skip conflicts silently

    print(f"  Dashboard synced: {synced} timesheet entries")


# ═══════════════════════════════════════════════════════════════
# STEP 6b: CHECK FOR REPLY EDITS & RESEND
# ═══════════════════════════════════════════════════════════════

def check_reply_edits(period_label):
    """
    Check Gayathri's inbox (or Anand's in production) for replies
    to the timesheet email. Parse edit instructions and apply them.
    """
    # In test mode, check Gayathri's sent-to-self replies
    # In production, check Anand's replies
    check_mailbox = TEST_EMAIL if TEST_MODE else ANAND_EMAIL
    search_subject = urllib.parse.quote(f"Consolidated Timesheet")

    print(f"\n  Checking {check_mailbox} for replies...")

    # Get recent replies to our timesheet email
    url = (
        f"/users/{SENDER_EMAIL}/messages"
        f"?$search={search_subject}"
        f"&$top=10"
        f"&$select=id,subject,from,receivedDateTime,body,conversationId"
    )
    try:
        result = graph_get(url)
    except Exception as e:
        print(f"  Error fetching emails: {e}")
        return None

    replies = []
    for m in result.get("value", []):
        subj = m.get("subject", "")
        frm = m.get("from", {}).get("emailAddress", {}).get("address", "")
        # Look for replies (Re: ...) that aren't from us
        if subj.startswith("Re:") or subj.startswith("RE:"):
            if period_label in subj or "Consolidated" in subj:
                replies.append(m)

    if not replies:
        print("  No replies found yet.")
        return None

    print(f"\n  Found {len(replies)} reply(ies):\n")
    for i, r in enumerate(replies):
        dt = r.get("receivedDateTime", "")[:16]
        frm = r.get("from", {}).get("emailAddress", {}).get("address", "")
        subj = r.get("subject", "")
        body = r.get("body", {}).get("content", "")
        # Clean body for display
        body_text = body.replace("\r\n", "\n").replace("<br>", "\n").replace("<br/>", "\n")
        # Strip HTML tags roughly
        import re
        body_text = re.sub(r'<[^>]+>', '', body_text)
        body_text = body_text.strip()[:1000]
        print(f"  [{i+1}] {dt} from {frm}")
        print(f"      Subject: {subj}")
        print(f"      Body:\n      {body_text[:500]}")
        print()

    return replies


def apply_edits(period_label, edits):
    """
    Apply edits to the timesheet data and regenerate Excel.
    edits: list of dicts like {"name": "Aravindh", "group": "D365_FO", "hours": 80}
    or {"name": "Aravindh", "group": "D365_FO", "action": "remove"}
    """
    period_dir = os.path.join(OUTPUT_DIR, period_label)

    # Load current data from the master file
    master_path = os.path.join(period_dir, f"Consolidated_VCC_ALL_{period_label}.xlsx")
    if not os.path.exists(master_path):
        print(f"  ERROR: {master_path} not found")
        return False

    wb = openpyxl.load_workbook(master_path)

    for edit in edits:
        name = edit.get("name", "")
        group = edit.get("group", "")
        action = edit.get("action", "update")

        # Find the sheet
        sheet_name = group[:31]
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found, skipping {name}")
            continue

        ws = wb[sheet_name]
        found = False
        for row in range(2, ws.max_row + 1):
            cell_name = ws.cell(row=row, column=3).value
            if cell_name and cell_name.lower().strip() == name.lower().strip():
                if action == "remove":
                    ws.delete_rows(row)
                    print(f"  Removed {name} from {group}")
                else:
                    new_hours = edit.get("hours", 0)
                    old_hours = ws.cell(row=row, column=4).value
                    ws.cell(row=row, column=4, value=new_hours)
                    print(f"  Updated {name} in {group}: {old_hours}h -> {new_hours}h")
                found = True
                break

        if not found and action != "remove":
            # Add new row
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=new_row - 1)
            ws.cell(row=new_row, column=2, value=edit.get("project", group))
            ws.cell(row=new_row, column=3, value=name)
            ws.cell(row=new_row, column=4, value=edit.get("hours", 0))
            print(f"  Added {name} to {group}: {edit.get('hours', 0)}h")

    # Save updated master
    wb.save(master_path)
    print(f"\n  Updated: {master_path}")
    return True


def ping_resource(resource_name, question, period_label):
    """Send verification to a resource via both Teams and email."""
    user = find_user_by_name(resource_name)
    if not user or not user.get("mail"):
        print(f"    Could not find {resource_name} in Azure AD")
        return False

    user_email = user["mail"]
    email_target = TEST_EMAIL if TEST_MODE else user_email

    # 1) Teams message
    teams_msg = (
        f"<b>Timesheet Verification — {period_label}</b><br><br>"
        f"Hi {resource_name},<br><br>"
        f"{question}<br><br>"
        f"Please reply to confirm. Thank you.<br>"
        f"— Gayathri"
    )
    print(f"    Teams → {resource_name} ({user_email})...", end=" ")
    teams_ok = send_teams_message(user_email, teams_msg)
    if teams_ok:
        print("SENT")
    else:
        print("FAILED (permission issue)")

    # 2) Email
    email_html = (
        f"<p>Hi {resource_name},</p>"
        f"<p>{question}</p>"
        f"<p>This is regarding your timesheet for period <b>{period_label}</b>. "
        f"Please reply to this email to confirm.</p>"
        f"<br><p>Thank you,<br>Gayathri</p>"
    )
    print(f"    Email → {resource_name} ({email_target})...", end=" ")
    try:
        send_email(
            from_email=SENDER_EMAIL,
            to_emails=[email_target],
            cc_emails=[],
            subject=f"Timesheet Verification - {resource_name} - {period_label}",
            html_body=email_html,
        )
        print("SENT")
    except Exception as e:
        print(f"FAILED: {e}")

    return True


# Keep old name as alias
ping_resource_on_teams = ping_resource


def interactive_edit_and_resend(period_label):
    """
    Interactive flow:
    1. Show Anand/reviewer's reply
    2. For each feedback item, choose action:
       - 'edit' → direct edit (change hours)
       - 'ask'  → ping the resource on Teams to verify
       - 'skip' → ignore this item for now
    3. After all inquiries resolved, apply edits and resend
    """
    replies = check_reply_edits(period_label)

    print("  ─────────────────────────────────────────────")
    print("  For each feedback item, choose an action:\n")
    print("  Commands:")
    print("    edit  name,group,new_hours   — Direct edit (e.g., edit Aravindh,D365_FO,80)")
    print("    ask   name,question          — Ping resource on Teams to verify")
    print("    remove name,group            — Remove resource from sheet")
    print("    done                         — Apply all edits and resend")
    print("    skip                         — Cancel everything")
    print("  ─────────────────────────────────────────────\n")

    edits = []
    inquiries_sent = []

    while True:
        line = input("  > ").strip()
        if not line:
            continue

        if line.lower() == "done":
            break
        if line.lower() == "skip":
            print("  Cancelled.")
            return

        parts = line.split(None, 1)  # Split on first space
        cmd = parts[0].lower()
        args = parts[1].strip() if len(parts) > 1 else ""

        if cmd == "edit":
            # edit name,group,hours
            fields = [f.strip() for f in args.split(",")]
            if len(fields) == 3:
                name, group, val = fields
                try:
                    edits.append({"name": name, "group": group, "hours": float(val)})
                    print(f"    Queued: {name} in {group} → {val}h")
                except ValueError:
                    print(f"    Invalid hours: {val}")
            else:
                print("    Format: edit name,group,new_hours")

        elif cmd == "ask":
            # ask name,question to ask them
            comma_pos = args.find(",")
            if comma_pos > 0:
                name = args[:comma_pos].strip()
                question = args[comma_pos + 1:].strip()
            else:
                name = args.strip()
                question = f"Did you work on the dates shown in your timesheet for period {period_label}? Please confirm."

            if name:
                ping_resource_on_teams(name, question, period_label)
                inquiries_sent.append(name)
            else:
                print("    Format: ask name,question (or just: ask name)")

        elif cmd == "remove":
            fields = [f.strip() for f in args.split(",")]
            if len(fields) == 2:
                name, group = fields
                edits.append({"name": name, "group": group, "action": "remove"})
                print(f"    Queued removal: {name} from {group}")
            else:
                print("    Format: remove name,group")

        else:
            print(f"    Unknown command: {cmd}")
            print("    Use: edit, ask, remove, done, skip")

    # Summary
    if inquiries_sent:
        print(f"\n  Teams messages sent to {len(inquiries_sent)} people: {', '.join(inquiries_sent)}")
        print("  Wait for their replies, then run 'edit' again to apply final changes.")

    if not edits:
        if inquiries_sent:
            print("  No edits to apply yet — waiting for confirmations.")
        else:
            print("  No edits to apply.")
        return

    print(f"\n  Applying {len(edits)} edit(s)...")
    apply_edits(period_label, edits)

    # Resend
    period_dir = os.path.join(OUTPUT_DIR, period_label)
    master_file = os.path.join(period_dir, f"Consolidated_VCC_ALL_{period_label}.xlsx")

    target = TEST_EMAIL if TEST_MODE else ANAND_EMAIL
    confirm = input(f"\n  Resend updated file to {target}? [y/N]: ").strip().lower()
    if confirm == "y":
        # Build change summary for the email
        change_lines = []
        for e in edits:
            if e.get("action") == "remove":
                change_lines.append(f"<li>Removed <b>{e['name']}</b> from {e['group']}</li>")
            else:
                change_lines.append(f"<li><b>{e['name']}</b> ({e['group']}): updated to {e['hours']}h</li>")
        changes_html = "<ul>" + "".join(change_lines) + "</ul>" if change_lines else ""

        inquiry_note = ""
        if inquiries_sent:
            inquiry_note = (
                f"<p><b>Pending verification:</b> Teams messages sent to "
                f"{', '.join(inquiries_sent)} — awaiting their confirmation.</p>"
            )

        html = f"""
        <p>Hi,</p>
        <p>Please find the <b>updated</b> Consolidated Timesheet for period <b>{period_label}</b>.</p>
        <p><b>Changes made:</b></p>
        {changes_html}
        {inquiry_note}
        <p>Kindly review and confirm if this is good to go.</p>
        <br>
        <p>Thank you and regards,<br>Gayathri</p>
        """
        send_email(
            from_email=SENDER_EMAIL,
            to_emails=[target],
            cc_emails=[],
            subject=f"[UPDATED] Consolidated Timesheet - VCC All Projects - {period_label}",
            html_body=html,
            attachments=[master_file],
        )
        print(f"  Resent updated file to {target}")
    else:
        print("  File updated locally but not resent.")


# ═══════════════════════════════════════════════════════════════
# MAIN CLI
# ═══════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    command = sys.argv[1].lower()
    period_label = sys.argv[2] if len(sys.argv) > 2 else "13AprTo-26Apr_2026"

    print(f"\n{'='*60}")
    print(f"VCC Timesheet Automation")
    print(f"Period: {period_label}")
    if TEST_MODE:
        print(f"*** TEST MODE — ALL emails go to {TEST_EMAIL} ***")
    else:
        print(f"*** PRODUCTION MODE — emails go to REAL clients ***")
    print(f"{'='*60}")

    if command == "check":
        # For demo, assume these have submitted (in real system, query from SharePoint/DB)
        submitted = input("  Enter comma-separated names of those who HAVE submitted\n  (or 'all' for everyone): ").strip()
        if submitted.lower() == "all":
            submitted_names = ALL_RESOURCES[:]
        else:
            submitted_names = [n.strip() for n in submitted.split(",")]
        check_completion(period_label, submitted_names)

    elif command == "generate":
        # Current period data (from user's input above)
        timesheet_data = {
            "IT_ALL": [
                {"name": "Senthilnathan", "project": "IT_HelpDesk", "hours": 80},
                {"name": "Vivekanandan", "project": "IT_HelpDesk", "hours": 96},
                {"name": "Aldrin", "project": "IT_Infra_8X8", "hours": 80},
                {"name": "Marcos", "project": "IT_Infra Onsite", "hours": 70},
                {"name": "Janani", "project": "IT Infra PMO", "hours": 80},
                {"name": "Janna", "project": "IT Infra PMO", "hours": 77},
                {"name": "Karthikeyan", "project": "IT_infra_Security", "hours": 80},
                {"name": "Mohammad", "project": "IT_infra_Security", "hours": 80},
                {"name": "Javal", "project": "Projects_financial_services", "hours": 80},
            ],
            "D365_FO": [
                {"name": "Manishkumar", "project": "D365_FO", "hours": 80},
                {"name": "Bradley", "project": "D365_FO", "hours": 80},
                {"name": "Kishore", "project": "D365_FO", "hours": 72},
                {"name": "Krupa", "project": "D365_FO", "hours": 85},
                {"name": "Andrea", "project": "D365_FO", "hours": 79},
                {"name": "Sankar", "project": "D365_FO", "hours": 80},
                {"name": "Anant", "project": "D365_FO", "hours": 72},
                {"name": "Shahul", "project": "D365_FO", "hours": 80},
                {"name": "Boyanarasimha", "project": "D365_FO", "hours": 80},
                {"name": "Srinivasan", "project": "D365_FO", "hours": 64},
                {"name": "Aravindh", "project": "D365_FO", "hours": 84},
                {"name": "Abhinandhan", "project": "D365_FO", "hours": 80},
                {"name": "Meenalochini", "project": "D365_FO", "hours": 80},
                {"name": "Keerthivasan", "project": "D365_FO", "hours": 77},
            ],
            "IT_Support_Savannah": [
                {"name": "Dhiraj", "project": "IT Savannah Support", "hours": 79},
                {"name": "Akhila", "project": "IT Savannah Support", "hours": 80},
            ],
            "QA_QC": [
                {"name": "Bindu", "project": "QA QC", "hours": 80},
                {"name": "Saritha", "project": "QA QC", "hours": 80},
                {"name": "Ganesh", "project": "QA QC", "hours": 80},
                {"name": "Manjari", "project": "QA QC", "hours": 80},
                {"name": "Mohammed", "project": "QA QC", "hours": 80},
                {"name": "Nishandhini", "project": "Salesforce", "hours": 80},
                {"name": "Aruldoss", "project": "Web B2B", "hours": 80},
                {"name": "Meenalochini", "project": "D365_FO", "hours": 80},
                {"name": "Shahul", "project": "D365_FO", "hours": 80},
                {"name": "Muthu", "project": "PartnerInsight", "hours": 80},
                {"name": "Madhavi", "project": "QA QC", "hours": 40},
                {"name": "Bhavesh", "project": "QA QC", "hours": 40},
                {"name": "Ashwath", "project": "QA Qc- Ecomm", "hours": 80},
            ],
            "JDE_EDI": [
                {"name": "Arul", "project": "JDE & EDI", "hours": 80},
                {"name": "Thomas", "project": "JDE & EDI", "hours": 80},
                {"name": "Bholeshankar", "project": "JDE & EDI", "hours": 80},
                {"name": "Nitinkumar", "project": "JDE & EDI", "hours": 72},
                {"name": "Chandan", "project": "JDE & EDI", "hours": 80},
                {"name": "Irfan", "project": "JDE & EDI", "hours": 40},
            ],
            "Enterprise_Integration": [
                {"name": "Pothiraja", "project": "Enterprise_Integration", "hours": 80},
                {"name": "Rafighafoor", "project": "Enterprise_Integration", "hours": 72},
            ],
            "Salesforce": [
                {"name": "Divya", "project": "Salesforce", "hours": 80},
                {"name": "Nishandhini", "project": "Salesforce", "hours": 80},
                {"name": "Naveenkumar", "project": "Salesforce", "hours": 80},
                {"name": "Karthiga", "project": "Salesforce", "hours": 80},
                {"name": "Sandhirasegaran", "project": "Salesforce", "hours": 72},
                {"name": "Hari", "project": "Salesforce", "hours": 80},
                {"name": "Swaminathan", "project": "Salesforce", "hours": 80},
                {"name": "Mohammed", "project": "Salesforce", "hours": 80},
            ],
            "PartnerInsight": [
                {"name": "Muthu", "project": "PartnerInsight", "hours": 112},
                {"name": "Sasikumar", "project": "PartnerInsight", "hours": 148},
                {"name": "Mahesh", "project": "PartnerInsight", "hours": 112},
                {"name": "Balaji", "project": "PartnerInsight", "hours": 148},
                {"name": "Kiran", "project": "PartnerInsight", "hours": 112},
                {"name": "Nageshwar", "project": "PartnerInsight", "hours": 140},
                {"name": "Gayathri", "project": "PartnerInsight", "hours": 80},
                {"name": "Humera", "project": "PartnerInsight", "hours": 80},
                {"name": "Zamir", "project": "PartnerInsight", "hours": 80},
            ],
            "Web_B2B": [
                {"name": "Aruldoss", "project": "Web B2B", "hours": 80},
                {"name": "Jagadesh", "project": "Web B2B", "hours": 80},
                {"name": "Sathishraj", "project": "Web B2B", "hours": 80},
                {"name": "Vimal", "project": "Web B2B", "hours": 80},
                {"name": "Anand S", "project": "Web B2B", "hours": 87},
            ],
        }

        # Auto-adjust for leaves before generating
        print("\n  Checking for leave records...")
        timesheet_data, leave_adjustments = adjust_timesheet_for_leaves(period_label, timesheet_data)
        if leave_adjustments:
            print(f"  {len(leave_adjustments)} leave adjustment(s) applied.\n")
        else:
            print("  No leave adjustments needed.\n")

        print("  Generating Excel files...\n")
        files, period_dir = generate_excel(period_label, timesheet_data)
        print(f"\n  Done! {len(files)} files in: {period_dir}")

    elif command == "send-anand":
        period_dir = os.path.join(OUTPUT_DIR, period_label)
        if not os.path.exists(period_dir):
            print(f"  ERROR: No files found for {period_label}. Run 'generate' first.")
            sys.exit(1)
        confirm = input(f"  Send timesheets to Anand ({ANAND_EMAIL})? [y/N]: ").strip().lower()
        if confirm == "y":
            send_to_anand(period_dir, period_label)
        else:
            print("  Cancelled.")

    elif command == "send-final":
        period_dir = os.path.join(OUTPUT_DIR, period_label)
        if not os.path.exists(period_dir):
            print(f"  ERROR: No files found for {period_label}. Run 'generate' first.")
            sys.exit(1)
        confirm = input(f"  Send FINAL approved timesheets to Gayathri ({GAYATHRI_EMAIL})? [y/N]: ").strip().lower()
        if confirm == "y":
            send_final_to_gayathri(period_dir, period_label)
        else:
            print("  Cancelled.")

    elif command == "edit":
        # Check for reply feedback and apply edits interactively
        interactive_edit_and_resend(period_label)

    elif command == "ai-review":
        # AI-powered review: parse Anand's email reply, apply edits, ping resources
        try:
            import ai_review
        except ImportError:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import ai_review
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        ai_review.run_ai_review(period_label, this_module)

    elif command == "ai-resolve":
        # Check Teams replies and resolve pending AI review items
        try:
            import ai_review
        except ImportError:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import ai_review
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        ai_review.resolve_pending(period_label, this_module)

    elif command == "leave-scan":
        # Scan leave tracker mailbox for new leave notifications
        new_leaves = scan_leave_emails(period_label)
        if new_leaves:
            print(f"\n  {len(new_leaves)} new leave(s) found.")
            # Check for excessive leaves and notify managers
            notifications = check_excessive_leaves_and_notify(period_label)
            if notifications:
                print(f"\n  {len(notifications)} manager notification(s) sent.")

    elif command == "leave-status":
        # Show leave summary
        leave_status(period_label)

    elif command == "leave-adjust":
        # Adjust timesheet hours based on recorded leaves
        period_dir = os.path.join(OUTPUT_DIR, period_label)
        if not os.path.exists(period_dir):
            print(f"  ERROR: No files for {period_label}. Run 'generate' first.")
            sys.exit(1)

        leaves = get_leaves_for_period(period_label)
        if not leaves:
            print("  No leaves recorded for this period. Nothing to adjust.")
        else:
            print(f"\n  Found {len(leaves)} leave record(s) for {period_label}")

            # Build timesheet data from current config
            timesheet_data = {}
            for group_key, config in TIMESHEET_CONFIG.items():
                timesheet_data[group_key] = []
                for r in config["resources"]:
                    timesheet_data[group_key].append({**r, "hours": r.get("hours", 80)})

            # Load actual hours from existing master file
            master_path = os.path.join(period_dir, f"Consolidated_VCC_ALL_{period_label}.xlsx")
            if os.path.exists(master_path):
                wb = openpyxl.load_workbook(master_path)
                for group_key in TIMESHEET_CONFIG:
                    sheet_name = group_key[:31]
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in range(5, ws.max_row + 1):  # Data starts after header rows
                            name_cell = ws.cell(row=row, column=3).value
                            hours_cell = ws.cell(row=row, column=4).value
                            if name_cell and hours_cell:
                                for r in timesheet_data.get(group_key, []):
                                    if r["name"].lower() == str(name_cell).lower().strip():
                                        r["hours"] = hours_cell

            # Apply leave adjustments
            adjusted_data, adjustments = adjust_timesheet_for_leaves(period_label, timesheet_data)

            if adjustments:
                print(f"\n  Regenerating timesheets with leave adjustments...")
                files, _ = generate_excel(period_label, adjusted_data)
                print(f"\n  {len(adjustments)} adjustment(s) applied. {len(files)} files regenerated.")

    elif command == "status":
        period_dir = os.path.join(OUTPUT_DIR, period_label)
        print(f"\n  Period: {period_label}")
        if os.path.exists(period_dir):
            files = [f for f in os.listdir(period_dir) if f.endswith(".xlsx")]
            print(f"  Generated files: {len(files)}")
            for f in sorted(files):
                size = os.path.getsize(os.path.join(period_dir, f))
                print(f"    {f} ({size/1024:.1f} KB)")
        else:
            print("  No files generated yet.")

        # Also show leave info
        leave_status(period_label)

    # ── HIRING COMMANDS ──────────────────────────────────────
    elif command == "hire-watch":
        interval = 5  # default 5 minutes
        if len(sys.argv) >= 4:
            try:
                interval = int(sys.argv[3])
            except ValueError:
                pass
        import hiring_automation
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        hiring_automation.run_hire_watch(this_module, interval_minutes=interval)

    elif command == "hire-scan":
        import hiring_automation
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        hiring_automation.run_hire_scan(this_module)

    elif command == "hire-jd":
        if len(sys.argv) < 4:
            print("  Usage: python timesheet_automation.py hire-jd <requirement_id>")
            sys.exit(1)
        req_id = sys.argv[3]
        import hiring_automation
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        hiring_automation.run_hire_jd(req_id, this_module)

    elif command == "hire-profiles":
        import hiring_automation
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        hiring_automation.run_hire_profiles(this_module)

    elif command == "hire-status":
        import hiring_automation
        hiring_automation.run_hire_status()

    elif command == "hire-onboard":
        if len(sys.argv) < 4:
            print("  Usage: python timesheet_automation.py hire-onboard <candidate_name>")
            sys.exit(1)
        name = " ".join(sys.argv[3:])
        import hiring_automation
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        hiring_automation.run_hire_onboard(name, this_module)

    elif command == "hire-update":
        if len(sys.argv) < 5:
            print("  Usage: python timesheet_automation.py hire-update <profile_id> <status>")
            print("  Statuses: received, shortlisted, interview_scheduled, interviewed,")
            print("            selected, rejected, offer_sent, offer_accepted, onboarding, joined, bench")
            sys.exit(1)
        import hiring_automation
        hiring_automation.run_hire_update(sys.argv[3], sys.argv[4])

    elif command == "hire-bench":
        import hiring_automation
        import importlib
        this_module = importlib.import_module("timesheet_automation")
        hiring_automation.run_hire_bench(this_module)

    else:
        print(f"  Unknown command: {command}")
        print("  Available: check, generate, send-anand, send-final, edit, ai-review, ai-resolve,")
        print("             leave-scan, leave-status, leave-adjust, status,")
        print("             hire-watch, hire-scan, hire-jd, hire-profiles, hire-status, hire-onboard, hire-update, hire-bench")


if __name__ == "__main__":
    main()
