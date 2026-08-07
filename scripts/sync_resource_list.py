"""
Sync VCC D4 April 2026 Consolidated Resources List → Supabase `users` table.
This Excel is the SINGLE SOURCE OF TRUTH for all employee/resource data.

Usage:
  python sync_resource_list.py                # Dry run (shows changes)
  python sync_resource_list.py --apply        # Apply changes to Supabase

It will:
1. Read all 69 resources from the consolidated Excel
2. Match to existing users in Supabase by name (fuzzy match)
3. Update ALL fields (project, role, email, location, joining date, etc.)
4. Create new users for unmatched resources
5. Deactivate users NOT in the list
6. Update sow_resources names to match
"""

import os
import sys
import io
import openpyxl
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Supabase connection — from the environment, never hardcoded.
#   export SUPABASE_URL=https://<ref>.supabase.co
#   export SUPABASE_KEY=<anon or service-role key>
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    sys.exit("ERROR: SUPABASE_URL and SUPABASE_KEY must be set in the environment.")

try:
    from supabase import create_client
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
except ImportError:
    print("ERROR: pip install supabase")
    sys.exit(1)

# Path to master Excel
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "VCC D4_APRIL_2026_Resources_List_Consolidated.xlsx")
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = "D:/Velozity Global/VCC/VCC D4_APRIL_2026_Resources_List_Consolidated.xlsx"
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = "D:/Velozity Global/VCC/timesheet-app/VCC D4_APRIL_2026_Resources_List_Consolidated.xlsx"

# Project name normalization (Excel project → standard project name used in app)
PROJECT_MAP = {
    "D365 F&O": "VCC - D365 F&O",
    "D365 F&O - Onsite": "VCC - D365 F&O",
    "D365 F&O Functional Onsite": "VCC - D365 F&O",
    "eComm BA - Onsite": "VCC - Web B2B",
    "eComm Support - Onsite": "VCC - Web B2B",
    "Enterprise Integration-Onsite": "VCC - Enterprise Integration",
    "Enterprise Integration": "VCC - Enterprise Integration",
    "Finance Assitant": "VCC - IT ALL",
    "IT Infra - Savannah": "VCC - IT Support Savannah",
    "IT Infra 8x8": "VCC - IT ALL",
    "IT Infra Help Desk": "VCC - IT ALL",
    "IT Infra Helpdesk Incident Manager - Onsite": "VCC - IT ALL",
    "IT Infra Helpdesk Onsite": "VCC - IT ALL",
    "IT Infra PMO": "VCC - IT ALL",
    "IT Infra Security": "VCC - IT ALL",
    "JDE EDI": "VCC - JDE EDI",
    "JDE EDI Onsite": "VCC - JDE EDI",
    "JDE onsite": "VCC - JDE EDI",
    "Partner Insight": "VCC - PartnerInsight",
    "Partner Insight & Web Support Onsite": "VCC - PartnerInsight",
    "Partner Insight Support Onsite": "VCC - PartnerInsight",
    "QA QC": "VCC - QA QC",
    "QA QC - Manger - Onsite": "VCC - QA QC",
    "QA QC - Performance": "VCC - QA QC",
    "QA QC- Ecommerce": "VCC - QA QC",
    "QA tester with E-commerce": "VCC - QA QC",
    "Salesforce": "VCC - Salesforce",
    "Salesforce BA": "VCC - Salesforce",
    "Web B2B": "VCC - Web B2B",
    "Delivery - IT": "VCC - Delivery",
}

# Known roles that should be manager/consolidator in the app
MANAGER_NAMES = {
    "anand kumar pandy": "manager",
    "krupa sarkar vyas": "manager",
    "bindu marella": "manager",
    "sandhisegaran munisami": "manager",
    "kishan vasant": "consolidator",
}

# Manual name mappings: Excel full name (lowercase) → DB name (for tricky mismatches)
NAME_OVERRIDES = {
    "aran thangaraj": "Thangaraj Aran",       # name order different in DB
    "raja pothi": "Pothiraja A",              # different format in DB
    "tilak gunasekaran": "Tilak Gunasekaran",   # full name
    "akash priyadharshan p": None,            # new person
}


def normalize_name(name):
    """Normalize name for matching."""
    return name.lower().strip().replace("  ", " ")


def fuzzy_match(excel_name, db_users):
    """Try to match an Excel name to a database user."""
    norm = normalize_name(excel_name)

    # Check override mapping first
    if norm in NAME_OVERRIDES:
        override_target = NAME_OVERRIDES[norm]
        if override_target is None:
            return None  # Explicitly marked as new/unmatched
        for u in db_users:
            if normalize_name(u["name"]) == normalize_name(override_target):
                return u

    # Exact match
    for u in db_users:
        if normalize_name(u["name"]) == norm:
            return u

    # Partial match (first + last name)
    parts = norm.split()
    for u in db_users:
        db_parts = normalize_name(u["name"]).split()
        if parts and db_parts and parts[0] == db_parts[0]:
            if len(parts) > 1 and len(db_parts) > 1:
                if parts[1][:3] == db_parts[1][:3]:
                    return u
            elif len(parts) == 1:
                return u

    # Try contains match
    for u in db_users:
        db_norm = normalize_name(u["name"])
        if norm in db_norm or db_norm in norm:
            return u

    return None


def generate_id(existing_ids):
    """Generate a unique employee ID."""
    max_id = 200000
    for eid in existing_ids:
        try:
            num = int(eid)
            if 200000 <= num < 300000 and num > max_id:
                max_id = num
        except (ValueError, TypeError):
            pass
    return str(max_id + 1)


def read_excel():
    """Read all resources from the consolidated Excel."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["D4_Resources"]

    resources = []
    for r in range(2, ws.max_row + 1):
        sno = ws.cell(r, 1).value
        if sno is None:
            continue

        full_name_raw = (ws.cell(r, 4).value or "").strip()
        # Remove "(D4 Contractor)" suffix
        full_name = full_name_raw.replace("(D4 Contractor)", "").replace("(D4 contractor)", "").strip()

        project_raw = (ws.cell(r, 5).value or "").strip()
        vcc_manager = (ws.cell(r, 6).value or "").strip()
        d4_manager = (ws.cell(r, 7).value or "").strip()
        role_title = (ws.cell(r, 8).value or "").strip()
        d4_email = (ws.cell(r, 9).value or "").strip()
        vcc_email = (ws.cell(r, 10).value or "").strip()
        location_raw = (ws.cell(r, 13).value or "").strip()
        joining_date = ws.cell(r, 14).value

        # Normalize project
        project = PROJECT_MAP.get(project_raw, f"VCC - {project_raw}")

        # Determine location type
        loc_type = "Offshore"
        if "onsite" in location_raw.lower() or "usa" in location_raw.lower():
            loc_type = "Onsite"

        # Determine app role (employee/manager/consolidator)
        app_role = MANAGER_NAMES.get(normalize_name(full_name), "employee")

        # Format joining date
        start_date = None
        if joining_date:
            try:
                start_date = joining_date.strftime("%Y-%m-%d") if hasattr(joining_date, 'strftime') else str(joining_date)[:10]
            except:
                pass

        resources.append({
            "sno": sno,
            "name": full_name,
            "project_raw": project_raw,
            "project": project,
            "designation": role_title if role_title else "Software Engineer",
            "d4_email": d4_email,
            "vcc_email": vcc_email,
            "d4_manager": d4_manager,
            "vcc_manager": vcc_manager,
            "location": location_raw,
            "location_type": loc_type,
            "start_date": start_date,
            "app_role": app_role,
        })

    return resources


def sync(apply=False):
    """Sync Excel resources to Supabase users table."""
    print("=" * 60)
    print("VCC D4 Consolidated Resources → Supabase Sync")
    print("=" * 60)

    # 1. Read Excel
    resources = read_excel()
    print(f"\nExcel: {len(resources)} resources")

    # 2. Fetch all current users from Supabase
    result = supabase.table("users").select("*").execute()
    db_users = result.data or []
    print(f"Supabase: {len(db_users)} users")

    existing_ids = [u["id"] for u in db_users]
    matched_db_ids = set()

    to_update = []
    to_create = []

    # 3. Match and compare
    print(f"\n{'─' * 60}")
    print("MATCHING RESOURCES:")
    print(f"{'─' * 60}")

    for res in resources:
        match = fuzzy_match(res["name"], db_users)

        if match:
            matched_db_ids.add(match["id"])
            changes = {}

            # Update project
            if match.get("project") != res["project"]:
                changes["project"] = res["project"]

            # Update designation
            if res["designation"] and match.get("designation") != res["designation"]:
                changes["designation"] = res["designation"]

            # Update email (use d4_email as primary)
            if res["d4_email"] and match.get("email") != res["d4_email"]:
                changes["email"] = res["d4_email"]

            # Update location type
            if res["location_type"] and match.get("location_type") != res["location_type"].lower():
                changes["location_type"] = res["location_type"].lower()

            # Update start_date if we have one and DB doesn't
            if res["start_date"] and not match.get("start_date"):
                changes["start_date"] = res["start_date"]

            # Update name if different (use Excel as source of truth)
            if match["name"] != res["name"]:
                changes["name"] = res["name"]

            # Update role if needed (but respect MANAGER_NAMES)
            if res["app_role"] != "employee" and match.get("role") != res["app_role"]:
                changes["role"] = res["app_role"]

            # Ensure active
            if match.get("is_active") != True:
                changes["is_active"] = True
                changes["is_archived"] = False
                changes["employee_status"] = "active"

            if changes:
                to_update.append({"id": match["id"], "name": match["name"], "new_name": res["name"], "changes": changes})
                print(f"  UPDATE: {res['name']} (matched → {match['name']})")
                for k, v in changes.items():
                    print(f"          {k}: {match.get(k)} → {v}")
            else:
                print(f"  OK:     {res['name']}")
        else:
            # New resource
            new_id = generate_id(existing_ids)
            existing_ids.append(new_id)
            to_create.append({
                "id": new_id,
                "name": res["name"],
                "role": res["app_role"],
                "designation": res["designation"],
                "project": res["project"],
                "email": res["d4_email"],
                "location_type": res["location_type"].lower(),
                "start_date": res["start_date"],
                "is_active": True,
                "is_archived": False,
                "employee_status": "active",
            })
            print(f"  NEW:    {res['name']} (ID: {new_id}, Project: {res['project']})")

    # 4. Find users to deactivate (in DB but not in Excel)
    # Skip ADM001 (system admin account)
    to_deactivate = []
    for user in db_users:
        if user["id"] not in matched_db_ids:
            if user["id"] == "ADM001":
                print(f"  KEEP:   {user['name']} (system admin)")
                continue
            if user.get("is_active") == True:
                to_deactivate.append(user)

    print(f"\n{'─' * 60}")
    print("SUMMARY:")
    print(f"{'─' * 60}")
    print(f"  Excel resources:      {len(resources)}")
    print(f"  Matched (no changes): {len(resources) - len(to_update) - len(to_create)}")
    print(f"  To update:            {len(to_update)}")
    print(f"  To create:            {len(to_create)}")
    print(f"  To deactivate:        {len(to_deactivate)}")

    if to_deactivate:
        print(f"\n  Will be DEACTIVATED (not in Excel):")
        for u in to_deactivate:
            print(f"    - {u['name']} ({u['id']}) [{u.get('project', '?')}] role={u.get('role')}")

    if not apply:
        print(f"\n{'─' * 60}")
        print("DRY RUN — no changes applied. Run with --apply to execute.")
        print(f"{'─' * 60}")
        return

    # 5. Apply changes
    print(f"\n{'─' * 60}")
    print("APPLYING CHANGES...")
    print(f"{'─' * 60}")

    for item in to_update:
        try:
            supabase.table("users").update({
                **item["changes"],
                "updated_at": datetime.now().isoformat()
            }).eq("id", item["id"]).execute()
            print(f"  ✓ Updated: {item.get('new_name', item['name'])}")
        except Exception as e:
            print(f"  ✗ Failed to update {item['name']}: {e}")

    for item in to_create:
        try:
            supabase.table("users").insert(item).execute()
            print(f"  ✓ Created: {item['name']} ({item['id']})")
        except Exception as e:
            print(f"  ✗ Failed to create {item['name']}: {e}")

    for user in to_deactivate:
        try:
            supabase.table("users").update({
                "is_active": False,
                "employee_status": "inactive",
                "updated_at": datetime.now().isoformat()
            }).eq("id", user["id"]).execute()
            print(f"  ✓ Deactivated: {user['name']} ({user['id']})")
        except Exception as e:
            print(f"  ✗ Failed to deactivate {user['name']}: {e}")

    # 6. Update sow_resources names to match new user names
    print(f"\n{'─' * 60}")
    print("SYNCING SOW RESOURCE NAMES...")
    print(f"{'─' * 60}")
    try:
        sow_result = supabase.table("sow_resources").select("id, name").execute()
        user_result = supabase.table("users").select("name").execute()
        user_names_set = {u["name"].lower().strip() for u in user_result.data}

        sow_updated = 0
        for sr in sow_result.data or []:
            sr_norm = sr["name"].lower().strip()
            if sr_norm not in user_names_set:
                # Try fuzzy match against updated users
                match = fuzzy_match(sr["name"], user_result.data)
                if match and match["name"] != sr["name"]:
                    supabase.table("sow_resources").update({"name": match["name"]}).eq("id", sr["id"]).execute()
                    print(f"  ✓ SOW: \"{sr['name']}\" → \"{match['name']}\"")
                    sow_updated += 1
        print(f"  SOW names updated: {sow_updated}")
    except Exception as e:
        print(f"  SOW sync error: {e}")

    print(f"\nDone! Users table synced with consolidated resource list ({len(resources)} resources).")


if __name__ == "__main__":
    apply = "--apply" in sys.argv
    sync(apply=apply)
